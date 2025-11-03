"""
🔬 TRACK ANALYSIS MODULE - TrackMate XML zu Diffusionsklassifikation
=======================================================================

Dieses Modul analysiert experimentelle Tracking-Daten aus TrackMate XML-Dateien
und klassifiziert Trajektorien-Segmente in 4 Diffusionsarten:
- Normal Diffusion
- Subdiffusion
- Confined Diffusion
- Superdiffusion

Features:
---------
- Multi-Scale Sliding Window Analyse (30-96 Frames)
- Random Forest Klassifikation mit trainiertem Modell
- MSD & Diffusionskoeffizienten (D) pro Segment
- Intelligente Glättung (min_segment_length)
- Vollständige Excel-Export (ein Sheet pro Track)
- Visualisierung (Pie Charts, Boxplots, Zeitreihen)

Version: 4.2 - ANGEPASST FÜR NEUE XML-STRUKTUR
Autor: Julian Lerch / Claude AI
"""

from __future__ import annotations

import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Tuple, NamedTuple
from dataclasses import dataclass, field
from collections import Counter, defaultdict
import json

import numpy as np
from joblib import load
from scipy import stats
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages


DiffusionLabel = str


# ============================================================================
# DATA STRUCTURES
# ============================================================================

class SpotData(NamedTuple):
    """Einzelner Spot aus TrackMate."""
    frame: int
    x: float
    y: float
    z: float
    t: float  # Zeit in Sekunden


@dataclass
class Track:
    """Eine Trajektorie mit allen Spots."""
    track_id: int
    spots: List[SpotData] = field(default_factory=list)
    name: str = ""

    def __len__(self) -> int:
        return len(self.spots)

    def get_positions(self) -> np.ndarray:
        """Returns Nx3 array of [x, y, z] positions."""
        return np.array([[s.x, s.y, s.z] for s in self.spots])

    def get_frames(self) -> np.ndarray:
        """Returns array of frame numbers."""
        return np.array([s.frame for s in self.spots])


@dataclass
class ClassifiedSegment:
    """Ein klassifiziertes Trajektorien-Segment."""
    start_frame: int
    end_frame: int
    diffusion_type: DiffusionLabel
    probability: float
    msd_slope: float  # α exponent
    D_value: float    # Diffusionskoeffizient [µm²/s]
    segment_length: int


@dataclass
class TrackAnalysis:
    """Analyse-Ergebnisse für einen Track."""
    track_id: int
    track_length: int
    segments: List[ClassifiedSegment]
    frame_labels: Dict[int, DiffusionLabel]  # Frame → Label mapping
    diffusion_distribution: Dict[DiffusionLabel, int]  # Counts
    mean_D_per_type: Dict[DiffusionLabel, float]
    mean_alpha_per_type: Dict[DiffusionLabel, float]


@dataclass
class XMLAnalysisResult:
    """Gesamt-Ergebnis für eine XML-Datei."""
    xml_path: Path
    num_tracks: int
    total_frames: int
    mean_track_length: float
    median_track_length: float
    track_analyses: List[TrackAnalysis]

    # Aggregierte Statistiken über alle Tracks
    global_diffusion_distribution: Dict[DiffusionLabel, int]
    global_D_values_per_type: Dict[DiffusionLabel, List[float]]
    global_alpha_values_per_type: Dict[DiffusionLabel, List[float]]


# ============================================================================
# TRACKMATE XML PARSER - ANGEPASST FÜR NEUE STRUKTUR
# ============================================================================

class TrackMateXMLParser:
    """
    Parser für TrackMate XML-Dateien.
    
    NEUE STRUKTUR:
    <Tracks>
        <particle nSpots="...">
            <detection t="0" x="..." y="..." z="..."/>
            <detection t="1" x="..." y="..." z="..."/>
            ...
        </particle>
    </Tracks>
    """

    def __init__(self, xml_path: Path):
        self.xml_path = Path(xml_path)
        self.tree = None
        self.root = None
        self.tracks: List[Track] = []

    def parse(self) -> List[Track]:
        """Parse XML und extrahiere alle Tracks."""
        
        print(f"   Parsing {self.xml_path.name}...")
        
        self.tree = ET.parse(str(self.xml_path))
        self.root = self.tree.getroot()

        # Parse Tracks direkt aus <particle> Elementen
        self.tracks = self._parse_particles()

        print(f"   ✅ {len(self.tracks)} Tracks erfolgreich geparst")
        
        return self.tracks

    def _parse_particles(self) -> List[Track]:
        """
        Parse alle <particle> Elemente.
        
        Jedes <particle> ist ein Track mit <detection> Elementen.
        """
        tracks = []
        
        # Finde alle <particle> Elemente
        particles = self.root.findall(".//particle")
        
        if not particles:
            # Fallback: Root könnte direkt <Tracks> sein
            if self.root.tag == "Tracks":
                particles = self.root.findall("particle")
        
        print(f"   Gefundene Particles: {len(particles)}")
        
        for track_id, particle in enumerate(particles):
            # Anzahl der Spots in diesem Track
            n_spots = int(particle.get("nSpots", 0))
            
            # Erstelle neuen Track
            track = Track(
                track_id=track_id,
                name=f"Track_{track_id}"
            )
            
            # Parse alle <detection> Elemente
            detections = particle.findall("detection")
            
            for detection in detections:
                # Zeit ist jetzt 't' Attribut (kontinuierlich in Sekunden)
                t = float(detection.get("t", 0.0))
                
                # Frame-Nummer aus Zeit berechnen
                # Annahme: frameInterval ist bekannt (aus Root-Attribut oder 1.0)
                frame_interval = float(self.root.get("frameInterval", "1.0"))
                frame_num = int(t / frame_interval)
                
                x = float(detection.get("x", 0.0))
                y = float(detection.get("y", 0.0))
                z = float(detection.get("z", 0.0))
                
                spot = SpotData(
                    frame=frame_num,
                    x=x,
                    y=y,
                    z=z,
                    t=t
                )
                
                track.spots.append(spot)
            
            # Sortiere Spots nach Zeit
            track.spots.sort(key=lambda s: s.t)
            
            # Nur Tracks mit genug Frames behalten (min 30)
            if len(track) >= 30:
                tracks.append(track)
            elif len(track) > 0:
                print(f"   ⚠️  Track {track_id} zu kurz ({len(track)} Frames < 30) - übersprungen")
        
        return tracks

    def get_preview_stats(self) -> Dict:
        """Erstelle Preview-Statistiken."""
        if not self.tracks:
            return {
                "num_tracks": 0,
                "total_frames": 0,
                "mean_length": 0,
                "median_length": 0,
                "min_length": 0,
                "max_length": 0
            }

        lengths = [len(t) for t in self.tracks]

        return {
            "num_tracks": len(self.tracks),
            "total_frames": sum(lengths),
            "mean_length": np.mean(lengths),
            "median_length": np.median(lengths),
            "min_length": min(lengths),
            "max_length": max(lengths),
            "tracks_histogram": np.histogram(lengths, bins=10)
        }


# ============================================================================
# MULTI-SCALE SLIDING WINDOW ANALYZER
# ============================================================================

class MultiScaleWindowAnalyzer:
    """Multi-Scale Sliding Window Analyse mit RF-Klassifikation."""

    # Feature Names (gleiche wie im RF-Training!)
    FEATURE_NAMES = [
        "mean_step_xy", "std_step_xy", "median_step_xy", "mad_step_xy",
        "max_step_xy", "min_step_xy", "mean_step_z", "std_step_z",
        "msd_lag1", "msd_lag2", "msd_lag4", "msd_lag8", "msd_loglog_slope",
        "straightness", "confinement_radius", "radius_of_gyration",
        "gyration_asymmetry", "turning_angle_mean", "turning_angle_std",
        "step_p90", "step_p10", "bounding_box_area", "axial_range",
        "directional_persistence", "velocity_autocorr",
        "step_skewness", "step_kurtosis"
    ]

    def __init__(self, model_path: Path, window_sizes: List[int] = None):
        """
        Parameters:
        -----------
        model_path : Path
            Pfad zum trainierten RF-Modell (.joblib)
        window_sizes : List[int]
            Liste von Window-Größen [Frames]. Default: [30, 48, 64, 96]
        """
        # Lade trainiertes Modell
        model_data = load(str(model_path))
        self.rf_model = model_data["model"]
        self.feature_names_trained = model_data.get("feature_names", self.FEATURE_NAMES)

        config = model_data.get("config", {}) or {}

        trained_windows = config.get("window_sizes") or []
        if not trained_windows and config.get("window_size"):
            trained_windows = [config.get("window_size")]

        if window_sizes:
            derived_windows = [int(w) for w in window_sizes if int(w) > 0]
        elif trained_windows:
            derived_windows = [int(w) for w in trained_windows if int(w) > 0]
        else:
            derived_windows = [30, 48, 64, 96]

        derived_windows = sorted({max(3, int(abs(w))) for w in derived_windows})
        self.window_sizes = derived_windows

        self.fixed_step_size = config.get("step_size")
        self.step_size_fraction = config.get("step_size_fraction", None)
        if self.step_size_fraction is None or self.step_size_fraction <= 0:
            self.step_size_fraction = 0.5

        self.min_segment_length_factor = config.get("min_segment_length_factor", 0.75)
        if not self.min_segment_length_factor or self.min_segment_length_factor <= 0:
            self.min_segment_length_factor = 0.75

        self.min_majority_fraction = config.get("min_majority_fraction", 0.7)
        self.keep_probability_threshold = max(0.5, min(0.95, self.min_majority_fraction - 0.05))

        smallest_window = min(self.window_sizes) if self.window_sizes else 30
        self.min_segment_length = max(3, int(round(smallest_window * self.min_segment_length_factor)))

        self.class_labels = list(getattr(self.rf_model, "classes_", []))
        if not self.class_labels:
            self.class_labels = ["subdiffusion", "normal", "confined", "superdiffusion"]
        self.default_label = self.class_labels[0]

    def analyze_track(self, track: Track, frame_rate_hz: float = 20.0) -> TrackAnalysis:
        """Analysiere einen kompletten Track mit Multi-Scale Windows."""

        positions = track.get_positions()
        frames = track.get_frames()

        # 1. Multi-Scale Prediction
        frame_predictions = self._multi_scale_classification(positions, frames)
        frame_predictions = self._prepare_frame_probabilities(frame_predictions, len(frames))

        # 2. Glätte Predictions (min_segment_length)
        smoothed_labels, frame_label_probabilities = self._smooth_predictions(
            frame_predictions,
            len(frames)
        )

        # 3. Finde Segmente (zusammenhängende Label-Regionen)
        segments = self._extract_segments(
            positions,
            frames,
            smoothed_labels,
            frame_rate_hz,
            frame_label_probabilities,
        )

        # 4. Berechne Statistiken
        diffusion_distribution = Counter(smoothed_labels.values())

        # Mean D und α pro Typ
        D_per_type = defaultdict(list)
        alpha_per_type = defaultdict(list)
        for seg in segments:
            D_per_type[seg.diffusion_type].append(seg.D_value)
            alpha_per_type[seg.diffusion_type].append(seg.msd_slope)

        mean_D = {k: np.mean(v) if v else 0.0 for k, v in D_per_type.items()}
        mean_alpha = {k: np.mean(v) if v else 0.0 for k, v in alpha_per_type.items()}

        return TrackAnalysis(
            track_id=track.track_id,
            track_length=len(track),
            segments=segments,
            frame_labels=smoothed_labels,
            diffusion_distribution=dict(diffusion_distribution),
            mean_D_per_type=mean_D,
            mean_alpha_per_type=mean_alpha
        )

    def _step_for_window(self, window_size: int) -> int:
        if self.fixed_step_size and self.fixed_step_size > 0:
            return max(1, min(int(window_size), int(self.fixed_step_size)))
        fraction = self.step_size_fraction if self.step_size_fraction and self.step_size_fraction > 0 else 0.5
        return max(1, int(round(window_size * fraction)))

    def _prepare_frame_probabilities(
        self,
        frame_predictions: Dict[int, Dict[str, float]],
        total_frames: int,
    ) -> Dict[int, Dict[str, float]]:
        if not frame_predictions:
            return {i: {self.default_label: 1.0} for i in range(total_frames)}

        counts = Counter()
        for probs in frame_predictions.values():
            if probs:
                best_label = max(probs.items(), key=lambda x: x[1])[0]
                counts[best_label] += 1

        fallback_label = counts.most_common(1)[0][0] if counts else self.default_label

        normalized = {}
        for frame_idx in range(total_frames):
            probs = frame_predictions.get(frame_idx)
            if not probs:
                normalized[frame_idx] = {fallback_label: 1.0}
                continue

            clean = {label: max(0.0, float(value)) for label, value in probs.items() if value is not None}
            total = sum(clean.values())
            if total <= 0:
                normalized[frame_idx] = {fallback_label: 1.0}
                continue

            normalized_probs = {label: value / total for label, value in clean.items()}
            # ensure fallback exists for stability
            if fallback_label not in normalized_probs:
                normalized_probs[fallback_label] = 0.0
            normalized[frame_idx] = normalized_probs

        return normalized

    def _multi_scale_classification(self, positions: np.ndarray,
                                   frames: np.ndarray) -> Dict[int, Dict[str, float]]:
        """
        Multi-Scale Window Klassifikation.

        Returns:
        --------
        Dict[frame_idx, Dict[label, probability]]
            Für jeden Frame: Wahrscheinlichkeiten aller Labels (über alle Window-Größen aggregiert)
        """

        frame_votes = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))  # frame → label → [sum_prob, total_weight]

        # Iteriere über alle Window-Größen
        for window_size in self.window_sizes:
            if positions.shape[0] < window_size:
                continue

            step_size = self._step_for_window(window_size)
            weight = float(window_size)

            for start in range(0, len(positions) - window_size + 1, step_size):
                end = start + window_size
                window_pos = positions[start:end]
                window_frames = frames[start:end]

                # Berechne Features
                features = self._compute_features(window_pos)
                if features is None:
                    continue

                # Predict
                proba = self.rf_model.predict_proba([features])[0]
                predicted_classes = self.rf_model.classes_

                # Addiere Votes für alle Frames in diesem Window
                for frame_idx in range(start, end):
                    for label, prob in zip(predicted_classes, proba):
                        vote_entry = frame_votes[frame_idx][label]
                        vote_entry[0] += float(prob) * weight
                        vote_entry[1] += weight

        # Aggregiere: Mittlere Wahrscheinlichkeit über alle Windows
        frame_predictions = {}
        for frame_idx, label_probs in frame_votes.items():
            aggregated = {}
            for label, (prob_sum, total_weight) in label_probs.items():
                if total_weight > 0:
                    aggregated[label] = prob_sum / total_weight
            frame_predictions[frame_idx] = aggregated

        return frame_predictions

    def _smooth_predictions(
        self,
        frame_predictions: Dict[int, Dict[str, float]],
        total_frames: int,
    ) -> Tuple[Dict[int, DiffusionLabel], Dict[int, float]]:
        """
        Glätte Predictions mit min_segment_length.

        Entfernt kurze Segmente (<min_length) und ersetzt sie durch Nachbar-Labels.
        """

        # 1. Wähle wahrscheinlichstes Label pro Frame
        raw_labels = {}
        for frame_idx in range(total_frames):
            probs = frame_predictions.get(frame_idx, {})
            if probs:
                best_label = max(probs.items(), key=lambda x: x[1])[0]
            else:
                best_label = self.default_label
            raw_labels[frame_idx] = best_label

        # 2. Finde Segmente
        segments = []
        current_label = None
        current_start = None

        for frame_idx in range(total_frames):
            label = raw_labels.get(frame_idx, self.default_label)

            if label != current_label:
                if current_label is not None:
                    segments.append((current_start, frame_idx - 1, current_label))
                current_label = label
                current_start = frame_idx

        # Letztes Segment
        if current_label is not None:
            segments.append((current_start, max(raw_labels.keys()), current_label))

        # 3. Glätte: Ersetze kurze Segmente
        smoothed_segments = []
        for i, (start, end, label) in enumerate(segments):
            length = end - start + 1
            probs = [frame_predictions.get(idx, {}).get(label, 0.0) for idx in range(start, end + 1)]
            avg_prob = float(np.mean(probs)) if probs else 0.0

            if length < self.min_segment_length and avg_prob < self.keep_probability_threshold:
                prev_label = segments[i - 1][2] if i > 0 else None
                next_label = segments[i + 1][2] if i < len(segments) - 1 else None

                candidates = []
                if prev_label is not None:
                    prev_probs = [frame_predictions.get(idx, {}).get(prev_label, 0.0) for idx in range(start, end + 1)]
                    candidates.append((float(np.mean(prev_probs)) if prev_probs else 0.0, prev_label))
                if next_label is not None:
                    next_probs = [frame_predictions.get(idx, {}).get(next_label, 0.0) for idx in range(start, end + 1)]
                    candidates.append((float(np.mean(next_probs)) if next_probs else 0.0, next_label))

                if candidates:
                    best_label = max(candidates, key=lambda x: x[0])[1]
                elif avg_prob >= self.keep_probability_threshold:
                    best_label = label
                else:
                    best_label = self.default_label

                smoothed_segments.append((start, end, best_label))
            else:
                smoothed_segments.append((start, end, label))

        # 4. Merge zusammenhängende Segmente mit gleichem Label
        merged = []
        for start, end, label in smoothed_segments:
            if merged and merged[-1][2] == label:
                # Erweitere vorheriges Segment
                merged[-1] = (merged[-1][0], end, label)
            else:
                merged.append((start, end, label))

        # 5. Konvertiere zu frame → label mapping
        smoothed_labels = {}
        frame_confidences = {}
        for start, end, label in merged:
            for frame_idx in range(start, end + 1):
                smoothed_labels[frame_idx] = label
                frame_confidences[frame_idx] = frame_predictions.get(frame_idx, {}).get(label, 0.0)

        return smoothed_labels, frame_confidences

    def _extract_segments(self, positions: np.ndarray, frames: np.ndarray,
                         frame_labels: Dict[int, DiffusionLabel],
                         frame_rate_hz: float,
                         frame_label_probabilities: Dict[int, float]) -> List[ClassifiedSegment]:
        """Extrahiere klassifizierte Segmente mit MSD/D Berechnung."""

        segments = []
        current_label = None
        current_start = None
        current_indices = []

        export_min_length = max(6, self.min_segment_length // 2)

        for i, frame_idx in enumerate(frames):
            label = frame_labels.get(i, self.default_label)

            if label != current_label:
                # Speichere vorheriges Segment
                if current_label is not None and len(current_indices) >= export_min_length:
                    seg_positions = positions[current_indices]
                    msd_slope, D_value = self._compute_msd_and_D(seg_positions, frame_rate_hz)
                    probs = [frame_label_probabilities.get(idx, 0.0) for idx in current_indices]
                    avg_prob = float(np.mean(probs)) if probs else 0.0

                    segments.append(ClassifiedSegment(
                        start_frame=int(frames[current_start]),
                        end_frame=int(frames[current_indices[-1]]),
                        diffusion_type=current_label,
                        probability=avg_prob,
                        msd_slope=msd_slope,
                        D_value=D_value,
                        segment_length=len(current_indices)
                    ))

                # Neues Segment starten
                current_label = label
                current_start = i
                current_indices = [i]
            else:
                current_indices.append(i)

        # Letztes Segment
        if current_label is not None and len(current_indices) >= export_min_length:
            seg_positions = positions[current_indices]
            msd_slope, D_value = self._compute_msd_and_D(seg_positions, frame_rate_hz)
            probs = [frame_label_probabilities.get(idx, 0.0) for idx in current_indices]
            avg_prob = float(np.mean(probs)) if probs else 0.0

            segments.append(ClassifiedSegment(
                start_frame=int(frames[current_start]),
                end_frame=int(frames[current_indices[-1]]),
                diffusion_type=current_label,
                probability=avg_prob,
                msd_slope=msd_slope,
                D_value=D_value,
                segment_length=len(current_indices)
            ))

        return segments

    def _compute_msd_and_D(self, positions: np.ndarray,
                          frame_rate_hz: float) -> Tuple[float, float]:
        """
        Berechne MSD-Slope (α) und Diffusionskoeffizient D.

        Returns:
        --------
        (msd_slope, D_value)
            msd_slope: α Exponent aus log(MSD) vs log(lag)
            D_value: Diffusionskoeffizient [µm²/s]
        """

        if len(positions) < 10:
            return 0.0, 0.0

        dt = 1.0 / frame_rate_hz  # Zeit pro Frame [s]

        # Berechne MSD für verschiedene Lags
        max_lag = min(len(positions) // 3, 20)
        lags = range(1, max_lag)
        msds = []

        for lag in lags:
            displacements = positions[lag:] - positions[:-lag]
            squared_displacements = np.sum(displacements**2, axis=1)
            msd = np.mean(squared_displacements)
            msds.append(msd)

        if len(msds) < 3:
            return 0.0, 0.0

        # Log-Log Regression für α
        lags_arr = np.array(list(lags))
        msds_arr = np.array(msds)

        mask = msds_arr > 0
        if np.sum(mask) < 3:
            return 0.0, 0.0

        log_lags = np.log(lags_arr[mask] * dt)
        log_msds = np.log(msds_arr[mask])

        try:
            slope, intercept, r_value, p_value, std_err = stats.linregress(log_lags, log_msds)
            msd_slope = slope  # α Exponent

            # Berechne D aus MSD(lag=1)
            # MSD(t) = 6*D*t für 3D normale Diffusion
            # MSD(t) = 6*D*t^α für anomale Diffusion
            # → D = MSD(t) / (6*t^α)
            if len(msds) > 0:
                msd_1 = msds[0]
                t_1 = dt
                D_value = msd_1 / (6.0 * (t_1 ** msd_slope) + 1e-12)
            else:
                D_value = 0.0

            return float(msd_slope), float(D_value)

        except:
            return 0.0, 0.0

    def _compute_features(self, window_pos: np.ndarray) -> Optional[np.ndarray]:
        """
        Berechne 27 Features für ein Positions-Window.

        WICHTIG: Gleiche Berechnung wie in rf_trainer.py!
        """

        steps = np.diff(window_pos, axis=0)
        if steps.size == 0:
            return None

        step_xy = np.linalg.norm(steps[:, :2], axis=1)
        step_z = np.abs(steps[:, 2])

        mean_step_xy = float(np.mean(step_xy))
        std_step_xy = float(np.std(step_xy))
        median_step_xy = float(np.median(step_xy))
        mad_step_xy = float(np.median(np.abs(step_xy - median_step_xy)))
        max_step_xy = float(np.max(step_xy))
        min_step_xy = float(np.min(step_xy))

        def _msd(lag: int) -> float:
            if lag >= window_pos.shape[0]:
                return 0.0
            diff = window_pos[lag:] - window_pos[:-lag]
            return float(np.mean(np.sum(diff**2, axis=1)))

        straight_dist = np.linalg.norm(window_pos[-1, :2] - window_pos[0, :2])
        total_path = np.sum(step_xy) + 1e-9
        straightness = straight_dist / total_path

        xy_centered = window_pos[:, :2] - np.mean(window_pos[:, :2], axis=0)
        confinement = float(np.sqrt(np.mean(np.sum(xy_centered**2, axis=1))))

        # Gyration
        centered = window_pos - np.mean(window_pos, axis=0)
        cov = np.dot(centered.T, centered) / max(1, centered.shape[0])
        eigvals = np.sort(np.real(np.linalg.eigvalsh(cov)))
        radius_of_gyration = float(np.sqrt(np.sum(np.clip(eigvals, 0.0, None))))
        denom = float(np.sum(np.abs(eigvals))) + 1e-12
        gyration_asymmetry = float((eigvals[-1] - eigvals[0]) / denom) if denom > 0 else 0.0

        # Turning angles
        if steps.shape[0] >= 2:
            v1 = steps[:-1, :2]
            v2 = steps[1:, :2]
            norms = (np.linalg.norm(v1, axis=1) * np.linalg.norm(v2, axis=1)) + 1e-12
            dots = np.sum(v1 * v2, axis=1)
            cos_angles = np.clip(dots / norms, -1.0, 1.0)
            turning_angles = np.arccos(cos_angles)
            turning_angle_mean = float(np.mean(turning_angles))
            turning_angle_std = float(np.std(turning_angles))

            # Directional persistence
            directional_persistence = float(np.mean(cos_angles))
        else:
            turning_angle_mean = 0.0
            turning_angle_std = 0.0
            directional_persistence = 0.0

        # Velocity autocorr
        if step_xy.size >= 2:
            a = step_xy[:-1]
            b = step_xy[1:]
            if np.std(a) > 1e-9 and np.std(b) > 1e-9:
                velocity_autocorr = float(np.corrcoef(a, b)[0, 1])
            else:
                velocity_autocorr = 0.0
        else:
            velocity_autocorr = 0.0

        if not np.isfinite(velocity_autocorr):
            velocity_autocorr = 0.0

        msd_lag1 = _msd(1)
        msd_lag2 = _msd(2)
        msd_lag4 = _msd(4)
        msd_lag8 = _msd(8)

        # MSD slope
        lags_arr = np.array([1, 2, 4, 8], dtype=np.float32)
        msd_arr = np.array([msd_lag1, msd_lag2, msd_lag4, msd_lag8], dtype=np.float32)
        mask = msd_arr > 0
        if np.count_nonzero(mask) >= 2:
            log_lags = np.log(lags_arr[mask])
            log_msds = np.log(msd_arr[mask])
            x_mean = float(np.mean(log_lags))
            y_mean = float(np.mean(log_msds))
            denom = float(np.sum((log_lags - x_mean) ** 2))
            if denom > 1e-12:
                msd_slope = float(np.sum((log_lags - x_mean) * (log_msds - y_mean)) / denom)
            else:
                msd_slope = 0.0
        else:
            msd_slope = 0.0

        # Moments
        if std_step_xy > 1e-12:
            centered_steps = (step_xy - mean_step_xy) / std_step_xy
            step_skewness = float(np.mean(centered_steps**3))
            step_kurtosis = float(np.mean(centered_steps**4) - 3.0)
        else:
            step_skewness = 0.0
            step_kurtosis = 0.0

        bbox = window_pos[:, :2]
        bbox_area = float((bbox[:, 0].max() - bbox[:, 0].min()) *
                         (bbox[:, 1].max() - bbox[:, 1].min()))
        axial_range = float(window_pos[:, 2].max() - window_pos[:, 2].min())

        feature_values = np.array([
            mean_step_xy, std_step_xy, median_step_xy, mad_step_xy,
            max_step_xy, min_step_xy,
            float(np.mean(step_z)), float(np.std(step_z)),
            msd_lag1, msd_lag2, msd_lag4, msd_lag8, msd_slope,
            float(straightness), confinement,
            radius_of_gyration, gyration_asymmetry,
            turning_angle_mean, turning_angle_std,
            float(np.percentile(step_xy, 90)), float(np.percentile(step_xy, 10)),
            bbox_area, axial_range,
            directional_persistence, velocity_autocorr,
            step_skewness, step_kurtosis
        ], dtype=np.float32)

        if np.any(~np.isfinite(feature_values)):
            return None

        return feature_values


# ============================================================================
# EXPORTERS
# ============================================================================

class AnalysisExporter:
    """Export von Analyse-Ergebnissen zu Excel, CSV und Plots."""

    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True, parents=True)

    def export_to_excel(self, xml_result: XMLAnalysisResult, tracks: List[Track]):
        """
        Exportiere zu Excel: Ein Sheet pro Track mit Frame-by-Frame Labels.

        Parameters:
        -----------
        xml_result : XMLAnalysisResult
            Analyse-Ergebnisse
        tracks : List[Track]
            Original-Tracks mit Positions-Daten

        Requires: openpyxl
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("⚠️  openpyxl nicht installiert. Installiere mit: pip install openpyxl")
            return None

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Farbschema für Diffusionstypen
        color_map = {
            "normal": "90EE90",  # Light green
            "subdiffusion": "FFD700",  # Gold
            "confined": "FF6B6B",  # Red
            "superdiffusion": "87CEEB"  # Sky blue
        }

        # Erstelle Mapping: track_id → Track
        tracks_dict = {t.track_id: t for t in tracks}

        for track_analysis in xml_result.track_analyses:
            track = tracks_dict.get(track_analysis.track_id)
            if track is None:
                continue

            # Neues Sheet pro Track
            sheet_name = f"Track_{track_analysis.track_id}"[:31]  # Excel limit
            ws = wb.create_sheet(title=sheet_name)

            # Header
            ws.append(["Frame", "X", "Y", "Z", "Time", "Diffusion_Type"])
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Frame-by-Frame Daten
            for i, spot in enumerate(track.spots):
                label = track_analysis.frame_labels.get(i, "unknown")
                ws.append([
                    spot.frame,
                    f"{spot.x:.6f}",
                    f"{spot.y:.6f}",
                    f"{spot.z:.6f}",
                    f"{spot.t:.3f}",
                    label
                ])

            ws.append([])
            ws.append(["=== SEGMENT SUMMARY ==="])
            ws.append(["Start_Frame", "End_Frame", "Type", "Length", "D [µm²/s]", "Alpha"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

            for seg in track_analysis.segments:
                ws.append([
                    seg.start_frame,
                    seg.end_frame,
                    seg.diffusion_type,
                    seg.segment_length,
                    f"{seg.D_value:.6f}",
                    f"{seg.msd_slope:.3f}"
                ])

        # Summary Sheet
        ws_summary = wb.create_sheet(title="Summary", index=0)
        ws_summary.append(["XML File:", xml_result.xml_path.name])
        ws_summary.append(["Total Tracks:", xml_result.num_tracks])
        ws_summary.append(["Mean Track Length:", f"{xml_result.mean_track_length:.1f}"])
        ws_summary.append([])
        ws_summary.append(["Diffusion Type", "Count", "Percentage"])

        total = sum(xml_result.global_diffusion_distribution.values())
        for dtype, count in xml_result.global_diffusion_distribution.items():
            pct = 100.0 * count / total if total > 0 else 0.0
            ws_summary.append([dtype, count, f"{pct:.1f}%"])

        output_path = self.output_dir / f"{xml_result.xml_path.stem}_analysis.xlsx"
        wb.save(output_path)

        return output_path

    def export_statistics_csv(self, xml_result: XMLAnalysisResult) -> Path:
        """Exportiere Statistiken als CSV."""
        import csv

        output_path = self.output_dir / f"{xml_result.xml_path.stem}_statistics.csv"

        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(["Metric", "Value"])
            writer.writerow(["XML File", xml_result.xml_path.name])
            writer.writerow(["Num Tracks", xml_result.num_tracks])
            writer.writerow(["Mean Track Length", f"{xml_result.mean_track_length:.1f}"])
            writer.writerow([])

            # Diffusion Distribution
            writer.writerow(["Diffusion Type", "Count", "Percentage"])
            total = sum(xml_result.global_diffusion_distribution.values())
            for dtype, count in xml_result.global_diffusion_distribution.items():
                pct = 100.0 * count / total if total > 0 else 0.0
                writer.writerow([dtype, count, f"{pct:.2f}"])

            writer.writerow([])

            # D Values
            writer.writerow(["Diffusion Type", "Mean D [µm²/s]", "Std D", "N"])
            for dtype, values in xml_result.global_D_values_per_type.items():
                if values:
                    writer.writerow([
                        dtype,
                        f"{np.mean(values):.6f}",
                        f"{np.std(values):.6f}",
                        len(values)
                    ])

            writer.writerow([])

            # Alpha Values
            writer.writerow(["Diffusion Type", "Mean Alpha", "Std Alpha", "N"])
            for dtype, values in xml_result.global_alpha_values_per_type.items():
                if values:
                    writer.writerow([
                        dtype,
                        f"{np.mean(values):.3f}",
                        f"{np.std(values):.3f}",
                        len(values)
                    ])

        return output_path

    def create_visualizations(self, xml_result: XMLAnalysisResult) -> List[Path]:
        """Erstelle Visualisierungen (Pie Chart, Boxplots)."""

        plots = []

        # 1. Pie Chart: Diffusion Distribution
        fig, ax = plt.subplots(figsize=(8, 6))

        labels = list(xml_result.global_diffusion_distribution.keys())
        sizes = list(xml_result.global_diffusion_distribution.values())
        colors = ['#90EE90', '#FFD700', '#FF6B6B', '#87CEEB'][:len(labels)]

        ax.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors, startangle=90)
        ax.set_title(f"Diffusion Type Distribution\n{xml_result.xml_path.name}")

        pie_path = self.output_dir / f"{xml_result.xml_path.stem}_distribution.pdf"
        plt.savefig(pie_path, bbox_inches='tight')
        plt.close()
        plots.append(pie_path)

        # 2. Boxplot: D Values
        fig, ax = plt.subplots(figsize=(10, 6))

        d_data = []
        d_labels = []
        for dtype in sorted(xml_result.global_D_values_per_type.keys()):
            values = xml_result.global_D_values_per_type[dtype]
            if values:
                d_data.append(values)
                d_labels.append(dtype)

        if d_data:
            bp = ax.boxplot(d_data, labels=d_labels, patch_artist=True)
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)

            ax.set_ylabel("Diffusion Coefficient D [µm²/s]")
            ax.set_title(f"D Values per Diffusion Type\n{xml_result.xml_path.name}")
            ax.grid(axis='y', alpha=0.3)

            d_path = self.output_dir / f"{xml_result.xml_path.stem}_D_boxplot.pdf"
            plt.savefig(d_path, bbox_inches='tight')
            plt.close()
            plots.append(d_path)

        # 3. Boxplot: Alpha Values
        fig, ax = plt.subplots(figsize=(10, 6))

        alpha_data = []
        alpha_labels = []
        for dtype in sorted(xml_result.global_alpha_values_per_type.keys()):
            values = xml_result.global_alpha_values_per_type[dtype]
            if values:
                alpha_data.append(values)
                alpha_labels.append(dtype)

        if alpha_data:
            bp = ax.boxplot(alpha_data, labels=alpha_labels, patch_artist=True)
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)

            ax.set_ylabel("MSD Exponent α")
            ax.set_title(f"Alpha Values per Diffusion Type\n{xml_result.xml_path.name}")
            ax.axhline(y=1.0, color='r', linestyle='--', alpha=0.5, label='Normal (α=1)')
            ax.grid(axis='y', alpha=0.3)
            ax.legend()

            alpha_path = self.output_dir / f"{xml_result.xml_path.stem}_alpha_boxplot.pdf"
            plt.savefig(alpha_path, bbox_inches='tight')
            plt.close()
            plots.append(alpha_path)

        return plots


# ============================================================================
# MAIN ANALYSIS ORCHESTRATOR
# ============================================================================

class TrackAnalysisOrchestrator:
    """Haupt-Orchestrator für die komplette Analyse-Pipeline."""

    def __init__(self, model_path: Path):
        self.model_path = Path(model_path)

        if not self.model_path.exists():
            raise FileNotFoundError(f"RF-Modell nicht gefunden: {model_path}")

        self.analyzer = MultiScaleWindowAnalyzer(self.model_path)

    def analyze_xml(self, xml_path: Path, output_dir: Path = None,
                   frame_rate_hz: float = 20.0) -> XMLAnalysisResult:
        """
        Analysiere eine TrackMate XML-Datei komplett.

        Parameters:
        -----------
        xml_path : Path
            Pfad zur TrackMate XML-Datei
        output_dir : Path, optional
            Output-Verzeichnis. Default: neben XML-Datei
        frame_rate_hz : float
            Frame-Rate für MSD/D Berechnung

        Returns:
        --------
        XMLAnalysisResult
            Komplette Analyse-Ergebnisse
        """

        xml_path = Path(xml_path)
        if output_dir is None:
            output_dir = xml_path.parent / f"{xml_path.stem}_analysis"
        else:
            output_dir = Path(output_dir)

        output_dir.mkdir(exist_ok=True, parents=True)

        print(f"📊 Analysiere: {xml_path.name}")

        # 1. Parse XML
        print("   Parsing XML...")
        parser = TrackMateXMLParser(xml_path)
        tracks = parser.parse()
        preview = parser.get_preview_stats()

        print(f"   ✅ {len(tracks)} Tracks gefunden")
        print(f"   📏 Mean Length: {preview['mean_length']:.1f} Frames")

        # 2. Analysiere jeden Track
        print("   Klassifiziere Tracks...")
        track_analyses = []

        for i, track in enumerate(tracks):
            if (i + 1) % 10 == 0:
                print(f"      Track {i+1}/{len(tracks)}...")

            analysis = self.analyzer.analyze_track(track, frame_rate_hz)
            track_analyses.append(analysis)

        # 3. Aggregiere Ergebnisse
        print("   Aggregiere Statistiken...")

        global_diffusion_dist = Counter()
        global_D_values = defaultdict(list)
        global_alpha_values = defaultdict(list)

        for analysis in track_analyses:
            for dtype, count in analysis.diffusion_distribution.items():
                global_diffusion_dist[dtype] += count

            for seg in analysis.segments:
                global_D_values[seg.diffusion_type].append(seg.D_value)
                global_alpha_values[seg.diffusion_type].append(seg.msd_slope)

        result = XMLAnalysisResult(
            xml_path=xml_path,
            num_tracks=len(tracks),
            total_frames=int(preview['total_frames']),
            mean_track_length=float(preview['mean_length']),
            median_track_length=float(preview['median_length']),
            track_analyses=track_analyses,
            global_diffusion_distribution=dict(global_diffusion_dist),
            global_D_values_per_type=dict(global_D_values),
            global_alpha_values_per_type=dict(global_alpha_values)
        )

        # 4. Exportiere Ergebnisse
        print("   Exportiere Ergebnisse...")
        exporter = AnalysisExporter(output_dir)

        # Excel (mit Original-Tracks für Positions-Daten)
        excel_path = exporter.export_to_excel(result, tracks)
        if excel_path:
            print(f"   ✅ Excel: {excel_path.name}")

        # CSV
        csv_path = exporter.export_statistics_csv(result)
        print(f"   ✅ CSV: {csv_path.name}")

        # Plots
        plots = exporter.create_visualizations(result)
        for plot in plots:
            print(f"   ✅ Plot: {plot.name}")

        print(f"✅ Analyse abgeschlossen: {output_dir}")

        return result

    def batch_analyze_folder(self, folder_path: Path, output_root: Path = None,
                            frame_rate_hz: float = 20.0, recursive: bool = True) -> List[XMLAnalysisResult]:
        """
        Analysiere alle XML-Dateien in einem Ordner.

        Parameters:
        -----------
        folder_path : Path
            Ordner mit XML-Dateien
        output_root : Path, optional
            Root-Output-Verzeichnis. Default: folder_path / "analysis_results"
        frame_rate_hz : float
            Frame-Rate
        recursive : bool
            Rekursiv in Unterordnern suchen?

        Returns:
        --------
        List[XMLAnalysisResult]
            Ergebnisse für alle XMLs
        """

        folder_path = Path(folder_path)
        if output_root is None:
            output_root = folder_path / "analysis_results"
        else:
            output_root = Path(output_root)

        # Finde alle XML-Dateien
        if recursive:
            xml_files = list(folder_path.rglob("*.xml"))
        else:
            xml_files = list(folder_path.glob("*.xml"))

        print(f"🔍 Gefunden: {len(xml_files)} XML-Dateien")

        results = []
        for i, xml_path in enumerate(xml_files):
            print(f"\n{'='*70}")
            print(f"[{i+1}/{len(xml_files)}] {xml_path.name}")
            print(f"{'='*70}")

            # Output-Dir für diese XML
            relative_path = xml_path.relative_to(folder_path)
            output_dir = output_root / relative_path.parent / f"{xml_path.stem}_analysis"

            try:
                result = self.analyze_xml(xml_path, output_dir, frame_rate_hz)
                results.append(result)
            except Exception as e:
                print(f"❌ FEHLER bei {xml_path.name}: {e}")
                import traceback
                traceback.print_exc()

        print(f"\n{'='*70}")
        print(f"✅ Batch-Analyse abgeschlossen: {len(results)}/{len(xml_files)} erfolgreich")
        print(f"📁 Output: {output_root}")

        return results


# ============================================================================
# CLI INTERFACE
# ============================================================================

def main():
    """CLI für Track-Analyse."""
    import argparse

    parser = argparse.ArgumentParser(
        description='🔬 Track Analysis - TrackMate XML zu Diffusionsklassifikation'
    )

    parser.add_argument(
        'xml_or_folder',
        type=str,
        help='Pfad zu TrackMate XML-Datei oder Ordner'
    )

    parser.add_argument(
        '--model',
        type=str,
        required=True,
        help='Pfad zum trainierten RF-Modell (.joblib)'
    )

    parser.add_argument(
        '--output',
        type=str,
        default=None,
        help='Output-Verzeichnis (default: neben XML)'
    )

    parser.add_argument(
        '--frame-rate',
        type=float,
        default=20.0,
        help='Frame-Rate [Hz] (default: 20)'
    )

    parser.add_argument(
        '--batch',
        action='store_true',
        help='Batch-Modus: Analysiere alle XMLs in Ordner'
    )

    parser.add_argument(
        '--recursive',
        action='store_true',
        help='Rekursiv in Unterordnern suchen (mit --batch)'
    )

    args = parser.parse_args()

    # Erstelle Orchestrator
    orchestrator = TrackAnalysisOrchestrator(Path(args.model))

    # Single oder Batch
    if args.batch:
        orchestrator.batch_analyze_folder(
            Path(args.xml_or_folder),
            Path(args.output) if args.output else None,
            args.frame_rate,
            args.recursive
        )
    else:
        orchestrator.analyze_xml(
            Path(args.xml_or_folder),
            Path(args.output) if args.output else None,
            args.frame_rate
        )


if __name__ == "__main__":
    main()